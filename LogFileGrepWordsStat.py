'''

@author: m1r
'''

import re
import numpy
from os import listdir
from os.path import isfile, join
#import xlsxwriter
import openpyxl
import datetime
import shutil
from collections import namedtuple, OrderedDict
from namedlist import namedlist

name_of_output = ''
sbj_basis = 6

column_map = {
    'r_h_o_hw_f': 1,
    'r_l_o_hw_f': 10,
    'l_h_o_hw_f': 19,
    'l_l_o_hw_f': 28,
    'r_h_o_lw_f': 37,
    'r_l_o_lw_f': 46,
    'l_h_o_lw_f': 55,
    'l_l_o_lw_f': 64,
    'r_h_c_hw_f': 73,
    'r_l_c_hw_f': 78,
    'l_h_c_hw_f': 83,
    'l_l_c_hw_f': 88,
    'r_h_c_lw_f': 93,
    'r_l_c_lw_f': 98,
    'l_h_c_lw_f': 103,
    'l_l_c_lw_f': 108,
    'r_h_c_cw_f': 113,
    'r_l_c_cw_f': 118,
    'l_h_c_cw_f': 123,
    'l_l_c_cw_f': 128,
    'r_h_o_hw_t': 133,
    'r_l_o_hw_t': 135,
    'l_h_o_hw_t': 137,
    'l_l_o_hw_t': 139,
    'r_h_o_lw_t': 141,
    'r_l_o_lw_t': 143,
    'l_h_o_lw_t': 145,
    'l_l_o_lw_t': 147,
    'r_h_o_cw_t': 149,
    'r_l_o_cw_t': 151,
    'l_h_o_cw_t': 153,
    'l_l_o_cw_t': 155,
    'r_h_c_hw_t': 157,
    'r_l_c_hw_t': 159,
    'l_h_c_hw_t': 161,
    'l_l_c_hw_t': 163,
    'r_h_c_lw_t': 165,
    'r_l_c_lw_t': 167,
    'l_h_c_lw_t': 169,
    'l_l_c_lw_t': 171,
    'r_h_c_cw_t': 173,
    'r_l_c_cw_t': 175,
    'l_h_c_cw_t': 177,
    'l_l_c_cw_t': 179
    }

def LastEmptyRow(ws, column):
    global sbj_basis
    ret = sbj_basis
    while True:
        if ws.cell(row = ret, column = column).value is None:
            return ret 
        else:
            ret += 1

def FindLastLine(ws):
    #global sbj_bias
    
    return max(map(lambda i : LastEmptyRow(ws, i), range(1,180)))
    #for column in range(1:180):
        

def GetColumn(side, limb, series_type, current_category, is_errors):
    if is_errors:
        postf = 't'
    else:
        postf = 'f'
    
    if len(side) == 0 or len(limb) == 0 or len(series_type) == 0 or len(current_category) == 0:
        tttt = 0
        
    search_for = side[0] + '_' + limb[0] + '_' + series_type[0] + '_' + current_category[0] + 'w_' + postf
    
    if not search_for in column_map:
        return None
      
    return column_map[search_for]

def WriteSubj(wb, pth):
    ws = wb.active
    
    WordStats = namedlist('WordStats', ['rt', 'cat','numerr'])
    
    active_side = ''#'right'
    current_category = ''#'hand'
    
    think_time = 0
    move_time = 0
    
    flnm = pth.split('/')[-1]
    
    date = flnm.split('_')[-1]
    date = date[:-4]
    
    info = flnm.split('_')[:-1]
    
    if 'hand' in info:
        limb = 'hand'
    else:
        limb = 'leg'
        
    if 'colored' in info:
        series_type = 'colored'
    else:
        series_type = 'ordinal'
    
    fread = open(pth, 'rb').read()
    mytext = fread.decode('utf-16')
    #mytext = mytext.encode('ascii', 'ignore')
    #fwrite = open('./tmp.log', 'wb')
    #fwrite.write(mytext)
    
    #fread.close()
    #fwrite.close()
    
    def DropData(side, limb, series_type, stats_map):
        od = OrderedDict(sorted(stats_map.items()))
        
        for k, v in od.items():
            column = GetColumn(side, limb, series_type, v.cat, False)
            
            if column is not None:
                ws.cell(row = LastEmptyRow(ws, column), column = column).value = k
                
                row = LastEmptyRow(ws, column + 1)
                
                if series_type == "colored":
                    lim = 5
                else:
                    lim = 9
                for i in range(1, min(len(v.rt)+1, lim)):
                    # iterate over columns here
                    ws.cell(row = row, column = column + i).value = v.rt[i-1]
            
            column = GetColumn(side, limb, series_type, v.cat, True)
            row    = LastEmptyRow(ws, column)
            ws.cell(row = row, column = column  ).value = k
            ws.cell(row = row, column = column+1).value = v.numerr
            
        wb.save(name_of_output)
        return
    
    current_word = ''
    sbj_id = ''
    
    #f = open('./tmp.log','r')
    
    hash_map_word_stat = {}
    for line in mytext.splitlines():#f:
        if (line.find("Subject id:") != -1):
            sbj_id = line.replace("Subject id: ", "")
            
            global sbj_basis
            found = False
            for rw in range(1,sbj_basis):
                if ws.cell(row=rw, column=1).value == sbj_id:
                    found = True
                    break
                
            if not found:
                last_empty_row = FindLastLine(ws)
                sbj_basis = last_empty_row + 1
                ws.merge_cells('A'+str(last_empty_row)+':'+'FX'+str(last_empty_row))
                ws['A'+str(last_empty_row)] = sbj_id
                wb.save(name_of_output)

        if (line.find("Active hand:") != -1):
            DropData(active_side, limb, series_type, hash_map_word_stat)
            
            hash_map_word_stat = {}
            
            if (line.find("RIGHT") != -1):
                active_side = 'right'
            
            if (line.find("LEFT") != -1):
                active_side = 'left'
            continue
    
        beg = line.find("Displaying text:")
        if (beg != -1):
            if (line.find("HAND") != -1):
                current_category = 'hand'
                
            if (line.find("LEG") != -1):
                current_category = 'leg'
                
            if (line.find("COMMON") != -1):
                current_category = 'common'
                
            line = line.replace("Displaying text: ", "")
            end = line.find("of category:")
            current_word = line[beg:end]
            continue
    
        if (line.find("Thinked before action:") != -1):
            think_time = int(re.findall(r'\d+', line)[0])
            continue
            
        if (line.find("Finger movement taken:") != -1):
            move_time = int(re.findall(r'\d+', line)[0])
            
            if current_word in hash_map_word_stat:
                hash_map_word_stat[current_word].rt.append(move_time+think_time)
            else:
                hash_map_word_stat[current_word] = WordStats(rt = [move_time+think_time], cat = current_category, numerr = 0)
            
            current_category = ""
            current_word     = ""
            continue
        
        if (line.find("Trial failed") != -1 and current_category != "" and current_word != ""):
            if current_word in hash_map_word_stat:
                hash_map_word_stat[current_word].numerr += 1
            else:
                hash_map_word_stat[current_word] = WordStats(rt = [], cat = current_category, numerr = 1)
            
            current_category = ""
            current_word     = ""
            continue
        
        if line.find("Experiment finished") != -1:
            DropData(active_side, limb, series_type, hash_map_word_stat)
         
    #f.close()
    #fread.close()


out_path = 'C:/Users/m1r/Documents/EmbodiedCognitionExpirement/release/subjects/logs/aggregated/'
template_path = 'C:/Users/m1r/Documents/EmbodiedCognitionExpirement/release/subjects/logs/aggregated/Template.xlsx'
dir_path = 'C:/Users/m1r/Documents/EmbodiedCognitionExpirement/release/subjects/logs/aggregated/all_logs/'

name_of_output = out_path + 'ResultingWordsStats_' + datetime.datetime.now().isoformat().replace(":","_").replace(".","_") + '.xlsx'
shutil.copy2(template_path, name_of_output)

#workbook = xlsxwriter.Workbook(out_path + 'ResultingWordsStats_' + datetime.datetime.now().isoformat() + '.xlsx')
#worksheet = workbook.add_worksheet()

#worksheet.write('A1', 'Hello world')

workbook = openpyxl.load_workbook(name_of_output)

onlyfiles = [f for f in listdir(dir_path) if isfile(join(dir_path, f))]

for sbj_log in onlyfiles:
    #ExtractInfo(dir_path + sbj_log)
    WriteSubj(workbook, dir_path + sbj_log)
#pth = 'C:/Users/m1r/Documents/EmbodiedCognitionExpirement/release/subjects/vasili_ordinal_hand_12.09.2016.log'

workbook.save(name_of_output)

#workbook.close()        